from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count, Case, When, Value, IntegerField, F, ExpressionWrapper, FloatField
from django.utils import timezone
from django.contrib.auth.models import User
from datetime import datetime
from .models import (
    Expediente, EtapaExpediente, DocumentoExpediente, ComentarioEtapa, 
    RequisitoEtapa, HistorialExpediente, Departamento, NotaExpediente, AreaTipoExpediente
)
from .forms import ExpedienteForm
import json
import zipfile
import os
from django.conf import settings
# PDF libraries y conversión de documentos
import io
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as ReportLabImage
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from docx import Document
from openpyxl import load_workbook
from PIL import Image as PILImage


def get_demo_user():
    """Obtiene o crea un usuario demo para el sistema"""
    user, created = User.objects.get_or_create(
        username='demo_user',
        defaults={
            'first_name': 'Usuario',
            'last_name': 'Demo',
            'email': 'demo@sistema.com',
            'is_active': True
        }
    )
    return user


def get_current_user(request):
    """Obtiene el usuario actual o el usuario demo si no está autenticado"""
    if request.user.is_authenticated:
        return request.user
    return get_demo_user()


def lista_expedientes(request):
    """Vista para listar expedientes con filtros"""
    expedientes = Expediente.objects.all()
    
    # Filtros
    tipo = request.GET.get('tipo')
    departamento = request.GET.get('departamento')
    estado = request.GET.get('estado')
    busqueda = request.GET.get('q')
    numero_sima = request.GET.get('numero_sima')
    usuario = request.GET.get('usuario')
    
    if tipo:
        expedientes = expedientes.filter(tipo_expediente=tipo)
    
    if departamento:
        expedientes = expedientes.filter(departamento_id=departamento)
    
    if estado:
        expedientes = expedientes.filter(estado_actual=estado)
    
    if busqueda:
        expedientes = expedientes.filter(
            Q(titulo__icontains=busqueda) |
            Q(numero_expediente__icontains=busqueda) |
            Q(descripcion__icontains=busqueda) |
            Q(palabras_clave__icontains=busqueda)
        )
    
    if numero_sima:
        expedientes = expedientes.filter(numero_sima__icontains=numero_sima)
    
    if usuario:
        expedientes = expedientes.filter(creado_por_id=usuario)
    
    # Paginación
    paginator = Paginator(expedientes, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Datos para filtros
    departamentos = Departamento.objects.filter(activo=True)
    tipos = Expediente.TIPO_CHOICES
    estados = Expediente.ESTADO_CHOICES
    usuarios = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    
    context = {
        'page_obj': page_obj,
        'departamentos': departamentos,
        'tipos': tipos,
        'estados': estados,
        'usuarios': usuarios,
        'filtros_actuales': {
            'tipo': tipo,
            'departamento': departamento,
            'estado': estado,
            'busqueda': busqueda,
            'numero_sima': numero_sima,
            'usuario': usuario,
        }
    }
    
    return render(request, 'digitalizacion/expedientes/lista_expedientes.html', context)


def seleccionar_tipo_expediente(request):
    """Vista para seleccionar el tipo de expediente a crear"""
    tipos = [
        {
            'key': 'giro',
            'title': 'Crear por Giro',
            'description': 'Inicia el expediente según el giro (p.ej. Ferretería).',
            'icon': 'bi-grid-3x3-gap'
        },
        {
            'key': 'fuente',
            'title': 'Crear por Fuente de Financiamiento',
            'description': 'Propio municipal, Estatal o Federal.',
            'icon': 'bi-bank'
        },
        {
            'key': 'tipo_adquisicion',
            'title': 'Crear por Tipo de Adquisición',
            'description': 'Bienes, Servicios, Arrendamientos.',
            'icon': 'bi-basket'
        },
        {
            'key': 'monto',
            'title': 'Crear por Monto',
            'description': 'Compra directa, Concurso por invitación, Licitación o Adjudicación directa.',
            'icon': 'bi-currency-dollar'
        },
    ]
    
    context = {
        'tipos': tipos
    }
    
    return render(request, 'digitalizacion/expedientes/seleccionar_tipo.html', context)


def crear_expediente(request, tipo):
    """Vista para crear un expediente según el tipo seleccionado"""
    from digitalizacion.models import Departamento
    
    if request.method == 'POST':
        # Generar número de expediente automático
        año_actual = datetime.now().year
        ultimo_numero = Expediente.objects.filter(
            numero_expediente__startswith=f"EXP-{año_actual}"
        ).count() + 1
        numero_expediente = f"EXP-{año_actual}-{ultimo_numero:04d}"
        
        # Obtener usuario (autenticado o demo)
        usuario = request.user if request.user.is_authenticated else get_demo_user()
        
        # Obtener departamento
        departamento_id = request.POST.get('departamento')
        if not departamento_id:
            # Si no se proporciona departamento, usar el primero disponible
            departamento = Departamento.objects.filter(activo=True).first()
            if not departamento:
                messages.error(request, 'No hay departamentos disponibles. Contacte al administrador.')
                return redirect('expedientes:seleccionar_tipo')
            departamento_id = departamento.id
        
        # Crear expediente base
        expediente = Expediente.objects.create(
            numero_expediente=numero_expediente,
            titulo=request.POST.get('titulo', 'Nuevo Expediente'),
            descripcion=request.POST.get('descripcion', ''),
            tipo_expediente=tipo,
            departamento_id=departamento_id,
            fecha_expediente=timezone.now().date(),
            creado_por=usuario,
            giro=request.POST.get('giro') if tipo == 'giro' else None,
            fuente_financiamiento=request.POST.get('fuente_financiamiento') if tipo == 'fuente' else None,
            tipo_adquisicion=request.POST.get('tipo_adquisicion') if tipo == 'tipo_adquisicion' else None,
            modalidad_monto=request.POST.get('modalidad_monto') if tipo == 'monto' else None,
        )
        
        # Crear las 17 etapas del expediente
        crear_etapas_expediente(expediente)
        
        # Crear entrada en el historial
        HistorialExpediente.objects.create(
            expediente=expediente,
            usuario=usuario,
            accion='Creación',
            descripcion=f'Expediente {expediente.get_tipo_expediente_display()} creado',
            etapa_nueva='inicio'
        )
        
        messages.success(request, f'Expediente {expediente.numero_expediente} creado exitosamente.')
        return redirect('expedientes:detalle', pk=expediente.pk)
    
    # GET request - mostrar formulario
    tipo_info = {
        'giro': {
            'title': 'Crear por Giro',
            'fields': ['giro'],
            'options': {}
        },
        'fuente': {
            'title': 'Crear por Fuente de Financiamiento',
            'fields': ['fuente_financiamiento'],
            'options': {
                'fuente_financiamiento': Expediente.FUENTE_CHOICES
            }
        },
        'tipo_adquisicion': {
            'title': 'Crear por Tipo de Adquisición',
            'fields': ['tipo_adquisicion'],
            'options': {
                'tipo_adquisicion': Expediente.TIPO_ADQUISICION_CHOICES
            }
        },
        'monto': {
            'title': 'Crear por Monto',
            'fields': ['modalidad_monto'],
            'options': {
                'modalidad_monto': Expediente.MODALIDAD_MONTO_CHOICES
            }
        }
    }
    
    context = {
        'tipo': tipo,
        'tipo_info': tipo_info.get(tipo, {}),
        'departamentos': Departamento.objects.filter(activo=True),
    }
    
    return render(request, 'digitalizacion/expedientes/crear_expediente.html', context)


def crear_etapas_expediente(expediente):
    """Crea las 17 etapas para un expediente"""
    etapas = [
        'inicio',
        'solicitud_area',
        'cotizacion',
        'requisicion_sima',
        'suficiencia_presupuestal',
        'aprobacion_director',
        'aprobacion_secretario',
        'notificacion_compras',
        'valoracion_tipo',
        'adjudicacion',
        'formalizacion',
        'contrato',
        'recepcion_bien',
        'recepcion_facturacion',
        'generacion_evidencia',
        'envio_compras',
        'pago',
    ]
    
    for etapa in etapas:
        EtapaExpediente.objects.create(
            expediente=expediente,
            nombre_etapa=etapa
        )


def detalle_expediente(request, pk):
    """Vista para mostrar los detalles de un expediente con sus etapas"""
    expediente = get_object_or_404(Expediente, pk=pk)
    
    # Definir el orden correcto de las etapas según el esquema solicitado
    orden_etapas = [
        'inicio',
        'solicitud_area',
        'cotizacion',
        'requisicion_sima',
        'suficiencia_presupuestal',
        'aprobacion_director',
        'aprobacion_secretario',
        'notificacion_compras',
        'valoracion_tipo',
        'adjudicacion',
        'formalizacion',
        'contrato',
        'recepcion_bien',
        'recepcion_facturacion',
        'generacion_evidencia',
        'envio_compras',
        'pago',
    ]
    
    # Obtener etapas en el orden correcto
    etapas_dict = {etapa.nombre_etapa: etapa for etapa in expediente.etapas.all()}
    etapas = [etapas_dict[nombre_etapa] for nombre_etapa in orden_etapas if nombre_etapa in etapas_dict]
    
    historial = expediente.historial.all()[:10]
    
    # Obtener requisitos según el tipo de expediente
    requisitos_por_etapa = {}
    for etapa in etapas:
        requisitos = obtener_requisitos_etapa(expediente, etapa.nombre_etapa)
        documentos = expediente.documentos.filter(etapa=etapa.nombre_etapa)
        comentarios = expediente.comentarios.filter(etapa=etapa.nombre_etapa)[:5]
        
        # Crear lista de nombres de documentos para verificación en template
        documentos_nombres = list(documentos.values_list('nombre_documento', flat=True))
        
        requisitos_por_etapa[etapa.nombre_etapa] = {
            'etapa': etapa,
            'requisitos': requisitos,
            'documentos': documentos,
            'documentos_nombres': documentos_nombres,
            'comentarios': comentarios,
            'puede_completar': puede_completar_etapa(expediente, etapa.nombre_etapa),
        }
    
    # Obtener el último documento subido
    ultimo_documento = expediente.documentos.order_by('-fecha_subida').first()
    
    # Obtener la última actualización (fecha del último documento o fecha de actualización del expediente)
    ultima_actualizacion = ultimo_documento.fecha_subida if ultimo_documento else expediente.fecha_actualizacion
    
    # Si hay un último documento y la fecha es más reciente que la de actualización del expediente
    if ultimo_documento and ultimo_documento.fecha_subida > expediente.fecha_actualizacion:
        expediente.fecha_actualizacion = ultimo_documento.fecha_subida
        expediente.save(update_fields=['fecha_actualizacion'])
    
    context = {
        'expediente': expediente,
        'etapas': etapas,
        'requisitos_por_etapa': requisitos_por_etapa,
        'historial': historial,
        'progreso': expediente.get_progreso(),
        'ultima_actualizacion': ultima_actualizacion,
        'ultimo_documento': ultimo_documento,
    }
    
    return render(request, 'digitalizacion/expedientes/detalle_expediente.html', context)


def obtener_requisitos_etapa(expediente, nombre_etapa):
    """Obtiene los requisitos específicos para una etapa según el tipo de expediente"""
    # Obtener requisitos base
    requisitos_base = list(RequisitoEtapa.objects.filter(
        tipo_expediente=expediente.tipo_expediente,
        etapa=nombre_etapa
    ).order_by('orden'))
    
    # Agregar requisitos específicos según subtipo
    requisitos_subtipo = []
    if expediente.tipo_expediente == 'tipo_adquisicion' and expediente.tipo_adquisicion:
        requisitos_subtipo = list(RequisitoEtapa.objects.filter(
            tipo_expediente='tipo_adquisicion',
            subtipo=expediente.tipo_adquisicion,
            etapa=nombre_etapa
        ).order_by('orden'))
    
    elif expediente.tipo_expediente == 'monto' and expediente.modalidad_monto:
        requisitos_subtipo = list(RequisitoEtapa.objects.filter(
            tipo_expediente='monto',
            subtipo=expediente.modalidad_monto,
            etapa=nombre_etapa
        ).order_by('orden'))
    
    # Combinar y ordenar manualmente
    todos_requisitos = requisitos_base + requisitos_subtipo
    # Eliminar duplicados basándose en nombre_requisito
    requisitos_unicos = {}
    for req in todos_requisitos:
        key = f"{req.etapa}_{req.nombre_requisito}"
        if key not in requisitos_unicos:
            requisitos_unicos[key] = req
    
    return sorted(requisitos_unicos.values(), key=lambda x: x.orden)


def puede_completar_etapa(expediente, nombre_etapa):
    """Verifica si una etapa puede ser completada"""
    # Definir el orden correcto de las etapas
    etapas_orden = [
        'inicio',
        'solicitud_area',
        'cotizacion',
        'requisicion_sima',
        'suficiencia_presupuestal',
        'aprobacion_director',
        'aprobacion_secretario',
        'notificacion_compras',
        'valoracion_tipo',
        'adjudicacion',
        'formalizacion',
        'contrato',
        'recepcion_bien',
        'recepcion_facturacion',
        'generacion_evidencia',
        'envio_compras',
        'pago',
    ]
    
    try:
        indice_actual = etapas_orden.index(nombre_etapa)
        if indice_actual > 0:
            etapa_anterior = etapas_orden[indice_actual - 1]
            etapa_anterior_obj = expediente.etapas.filter(nombre_etapa=etapa_anterior).first()
            if not etapa_anterior_obj or not etapa_anterior_obj.completada:
                return False
    except ValueError:
        return False
    
    # Verificar que hay al menos un documento subido en esta etapa
    documentos_etapa = expediente.documentos.filter(etapa=nombre_etapa)
    if not documentos_etapa.exists():
        return False
    
    # Para efectos de demo, si hay documentos subidos, permitir completar
    # En una implementación más rigurosa, aquí se verificarían requisitos específicos
    
    # Opcional: Verificar requisitos específicos si están definidos en la base de datos
    requisitos_base = RequisitoEtapa.objects.filter(
        tipo_expediente=expediente.tipo_expediente,
        etapa=nombre_etapa,
        obligatorio=True
    )
    
    # Si no hay requisitos específicos definidos, permitir completar si hay documentos
    if not requisitos_base.exists():
        return True
    
    # Si hay requisitos definidos, verificar que al menos hay documentos 
    # (por ahora simplificado para que funcione el demo)
    return documentos_etapa.count() > 0


from django.utils import timezone

def subir_documento(request, expediente_id, etapa):
    """Vista para subir documentos a una etapa específica"""
    from django.core.cache import cache
    
    expediente = get_object_or_404(Expediente, pk=expediente_id)
    
    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        nombre_documento = request.POST.get('nombre_documento')
        descripcion = request.POST.get('descripcion', '')
        numero_sima = request.POST.get('numero_sima', '')
        
        if archivo and nombre_documento:
            # Obtener usuario (autenticado o demo)
            usuario = request.user if request.user.is_authenticated else get_demo_user()
            
            # Si es la etapa de requisición SIMA, guardar el número SIMA en el expediente
            if etapa == 'requisicion_sima' and numero_sima:
                expediente.numero_sima = numero_sima
                expediente.save()
                messages.success(request, f'Número SIMA {numero_sima} asignado al expediente')
            
            # Crear el documento
            documento = DocumentoExpediente.objects.create(
                expediente=expediente,
                etapa=etapa,
                nombre_documento=nombre_documento,
                archivo=archivo,
                descripcion=descripcion,
                subido_por=usuario
            )
            
            # Limpiar la caché de última actualización
            cache_key = f'expediente_{expediente.id}_ultima_actualizacion'
            cache.delete(cache_key)
            
            # Actualizar manualmente la fecha de actualización del expediente
            ahora = timezone.now()
            expediente.fecha_actualizacion = ahora
            expediente.save(update_fields=['fecha_actualizacion'])
            
            # Crear entrada en el historial
            HistorialExpediente.objects.create(
                expediente=expediente,
                usuario=usuario,
                accion='Documento subido',
                descripcion=f'Documento "{nombre_documento}" subido en etapa {etapa}',
                etapa_nueva=etapa
            )
            
            # Forzar la actualización de la caché con la nueva fecha
            expediente.get_ultima_actualizacion(use_cache=False)
            
            messages.success(request, f'Documento "{nombre_documento}" subido exitosamente.')
        else:
            messages.error(request, 'Archivo y nombre del documento son requeridos.')
    
    return redirect('detalle_expediente', pk=expediente_id)


def completar_etapa(request, expediente_id, etapa):
    """Vista para marcar una etapa como completada"""
    expediente = get_object_or_404(Expediente, pk=expediente_id)
    etapa_obj = get_object_or_404(EtapaExpediente, expediente=expediente, nombre_etapa=etapa)
    
    if request.method == 'POST':
        if puede_completar_etapa(expediente, etapa):
            # Obtener usuario (autenticado o demo)
            usuario = request.user if request.user.is_authenticated else get_demo_user()
            
            # Actualizar la etapa
            etapa_obj.completada = True
            etapa_obj.fecha_completada = timezone.now()
            etapa_obj.completada_por = usuario
            etapa_obj.notas = request.POST.get('notas', '')
            etapa_obj.save()
            
            # Actualizar la fecha de actualización del expediente
            expediente.fecha_actualizacion = timezone.now()
            expediente.save(update_fields=['fecha_actualizacion'])
            
            # Obtener la siguiente etapa en el orden correcto
            orden_etapas = [
                'inicio',
                'solicitud_area', 
                'cotizacion',
                'requisicion_sima',
                'suficiencia_presupuestal',
                'aprobacion_director',
                'aprobacion_secretario',
                'notificacion_compras',
                'valoracion_tipo',
                'adjudicacion',
                'formalizacion',
                'contrato',
                'recepcion_bien',
                'recepcion_facturacion',
                'generacion_evidencia',
                'envio_compras',
                'pago',
            ]
            
            try:
                indice_actual = orden_etapas.index(etapa)
                if indice_actual < len(orden_etapas) - 1:
                    # Avanzar a la siguiente etapa
                    siguiente_etapa = orden_etapas[indice_actual + 1]
                    expediente.estado_actual = siguiente_etapa
                else:
                    # Es la última etapa, marcar como completado
                    expediente.estado_actual = 'completado'
                # Actualizar la fecha de actualización
                expediente.fecha_actualizacion = timezone.now()
                expediente.save()
            except ValueError:
                # Si la etapa no está en la lista, mantener el estado actual
                pass
            
            # Crear entrada en el historial
            HistorialExpediente.objects.create(
                expediente=expediente,
                usuario=usuario,
                accion='Etapa completada',
                descripcion=f'Etapa "{etapa_obj.get_nombre_etapa_display()}" marcada como completada',
                etapa_anterior=etapa,
                etapa_nueva=expediente.estado_actual
            )
            
            # Mensaje con información de progreso
            progreso = expediente.get_progreso()
            if expediente.estado_actual == 'completado':
                messages.success(request, f'¡Expediente completado al 100%! Todas las etapas han sido finalizadas.')
            else:
                messages.success(request, f'Etapa "{etapa_obj.get_nombre_etapa_display()}" completada. Progreso actual: {progreso}%')
        else:
            messages.error(request, 'No se puede completar esta etapa. Verifique los requisitos obligatorios.')
    
    return redirect('expedientes:detalle', pk=expediente_id)


def agregar_comentario(request, expediente_id, etapa):
    """Vista para agregar comentarios a una etapa"""
    expediente = get_object_or_404(Expediente, pk=expediente_id)
    
    if request.method == 'POST':
        comentario_texto = request.POST.get('comentario')
        
        if comentario_texto:
            # Obtener usuario (autenticado o demo)
            usuario = request.user if request.user.is_authenticated else get_demo_user()
            
            ComentarioEtapa.objects.create(
                expediente=expediente,
                etapa=etapa,
                comentario=comentario_texto,
                usuario=usuario
            )
            
            messages.success(request, 'Comentario agregado exitosamente.')
        else:
            messages.error(request, 'El comentario no puede estar vacío.')
    
    return redirect('expedientes:detalle', pk=expediente_id)


def dashboard_expedientes(request):
    """Dashboard principal para expedientes"""
    # Estadísticas básicas
    total_expedientes = Expediente.objects.count()
    expedientes_en_proceso = Expediente.objects.exclude(estado_actual='completado').count()
    expedientes_completados = Expediente.objects.filter(estado_actual='completado').count()
    
    # Calcular porcentaje de eficiencia
    if total_expedientes > 0:
        porcentaje_eficiencia = round((expedientes_completados * 100) / total_expedientes)
    else:
        porcentaje_eficiencia = 0
    
    # Expedientes recientes
    expedientes_recientes = Expediente.objects.order_by('-fecha_creacion')[:5]
    
    # Expedientes por tipo
    expedientes_por_tipo = Expediente.objects.values('tipo_expediente').annotate(
        total=Count('tipo_expediente')
    ).order_by('tipo_expediente')
    
    # Expedientes por estado
    expedientes_por_estado = Expediente.objects.values('estado_actual').annotate(
        total=Count('estado_actual')
    ).order_by('estado_actual')
    
    context = {
        'total_expedientes': total_expedientes,
        'expedientes_en_proceso': expedientes_en_proceso,
        'expedientes_completados': expedientes_completados,
        'porcentaje_eficiencia': porcentaje_eficiencia,
        'expedientes_recientes': expedientes_recientes,
        'expedientes_por_tipo': expedientes_por_tipo,
        'expedientes_por_estado': expedientes_por_estado,
    }
    
    return render(request, 'digitalizacion/expedientes/dashboard.html', context)


@login_required
def rechazar_expediente(request, expediente_id):
    """Vista para rechazar un expediente"""
    expediente = get_object_or_404(Expediente, pk=expediente_id)
    
    if request.method == 'POST':
        motivo_rechazo = request.POST.get('motivo_rechazo')
        
        if motivo_rechazo:
            # Obtener usuario (autenticado o demo)
            usuario = request.user if request.user.is_authenticated else get_demo_user()
            
            # Rechazar el expediente usando el método del modelo
            expediente.rechazar_expediente(usuario, motivo_rechazo)
            
            messages.success(request, f'Expediente {expediente.numero_expediente} ha sido rechazado correctamente.')
            return redirect('expedientes:detalle', pk=expediente.pk)
        else:
            messages.error(request, 'Debe proporcionar un motivo para rechazar el expediente.')
    
    return redirect('expedientes:detalle', pk=expediente.pk)


def visualizador_expediente(request, expediente_id):
    """Vista para el visualizador colaborativo por áreas"""
    from .models import ComentarioArea, Notificacion
    
    expediente = get_object_or_404(Expediente, pk=expediente_id)
    
    # Orden de etapas definido
    orden_etapas = [
        'inicio', 'solicitud_area', 'cotizacion', 'requisicion_sima',
        'suficiencia_presupuestal', 'aprobacion_director', 'aprobacion_secretario',
        'notificacion_compras', 'valoracion_tipo', 'adjudicacion',
        'formalizacion', 'contrato', 'recepcion_bien', 'recepcion_facturacion',
        'generacion_evidencia', 'envio_compras', 'pago',
    ]
    
    # Nombres de etapas formateados
    nombres_etapas = {
        'inicio': 'INICIO',
        'solicitud_area': 'SOLICITUD DEL ÁREA',
        'cotizacion': 'COTIZACIÓN',
        'requisicion_sima': 'REQUISICIÓN SIMA',
        'suficiencia_presupuestal': 'SUFICIENCIA PRESUPUESTAL',
        'aprobacion_director': 'APROBACIÓN DIRECTOR ADMINISTRATIVO',
        'aprobacion_secretario': 'APROBACIÓN SECRETARIO',
        'notificacion_compras': 'NOTIFICACIÓN A COMPRAS MUNICIPALES',
        'valoracion_tipo': 'VALORACIÓN PARA TIPO DE ADQUISICIÓN',
        'adjudicacion': 'ADJUDICACIÓN',
        'formalizacion': 'FORMALIZACIÓN CON ORDEN DE COMPRA',
        'contrato': 'CONTRATO',
        'recepcion_bien': 'RECEPCIÓN DEL BIEN O SERVICIO',
        'recepcion_facturacion': 'RECEPCIÓN DE FACTURACIÓN',
        'generacion_evidencia': 'GENERACIÓN DE EVIDENCIA',
        'envio_compras': 'ENVÍO DE EXPEDIENTE A COMPRAS',
        'pago': 'PAGO',
    }
    
    # Obtener etapas del expediente
    etapas_expediente = EtapaExpediente.objects.filter(
        expediente=expediente
    ).order_by('nombre_etapa')
    
    # Organizar información por áreas/etapas
    areas_data = []
    for etapa_key in orden_etapas:
        # Buscar si existe esta etapa en el expediente
        etapa_obj = etapas_expediente.filter(nombre_etapa=etapa_key).first()
        
        if etapa_obj or True:  # Mostrar todas las etapas, tengan documentos o no
            # Documentos de esta etapa
            documentos = DocumentoExpediente.objects.filter(
                expediente=expediente,
                etapa=etapa_key
            ).order_by('fecha_subida')
            
            # Comentarios de esta área
            comentarios = ComentarioArea.objects.filter(
                expediente=expediente,
                etapa=etapa_key
            ).order_by('fecha_creacion')
            
            # Notas post-it de documentos en esta etapa
            notas_etapa = []
            for doc in documentos:
                notas_doc = NotaExpediente.objects.filter(
                    expediente=expediente,
                    documento_id=doc.id,
                    activa=True
                )
                notas_etapa.extend(notas_doc)
            
            # Preparar URLs absolutas para documentos
            for documento in documentos:
                if documento.archivo:
                    documento.url_absoluta = request.build_absolute_uri(documento.archivo.url)
            
            area_data = {
                'etapa_key': etapa_key,
                'etapa_obj': etapa_obj,
                'nombre_etapa': nombres_etapas.get(etapa_key, etapa_key.replace('_', ' ').title()),
                'documentos': documentos,
                'comentarios': comentarios,
                'notas': notas_etapa,
                'tiene_documentos': documentos.exists(),
                'completada': etapa_obj.fecha_completada is not None if etapa_obj else False,
            }
            areas_data.append(area_data)
    
    # Obtener notificaciones no leídas del usuario actual
    notificaciones_usuario = Notificacion.objects.filter(
        usuario=request.user,
        expediente=expediente,
        leida=False
    ).order_by('-fecha_creacion')
    
    # Obtener todos los usuarios para menciones
    usuarios_disponibles = User.objects.filter(is_active=True).exclude(id=request.user.id)
    
    # Calcular progreso
    etapas_con_documentos = [area for area in areas_data if area['tiene_documentos']]
    etapas_completadas = [area for area in areas_data if area['completada']]
    progreso = (len(etapas_completadas) / len(etapas_con_documentos) * 100) if etapas_con_documentos else 0
    
    context = {
        'expediente': expediente,
        'areas_data': areas_data,
        'notificaciones': notificaciones_usuario,
        'usuarios_disponibles': usuarios_disponibles,
        'colores_notas': NotaExpediente.COLORES_CHOICES,
        'total_etapas': len(etapas_con_documentos),
        'etapas_completadas': len(etapas_completadas),
        'progreso': round(progreso, 1),
    }
    
    return render(request, 'digitalizacion/expedientes/visualizador_colaborativo.html', context)


@login_required
def crear_nota(request, expediente_id):
    """Vista para crear una nueva nota en el expediente"""
    expediente = get_object_or_404(Expediente, pk=expediente_id)
    
    if request.method == 'POST':
        contenido = request.POST.get('contenido')
        color = request.POST.get('color', 'amarillo')
        documento_id = request.POST.get('documento_id')
        posicion_x = int(request.POST.get('posicion_x', 0))
        posicion_y = int(request.POST.get('posicion_y', 0))
        
        if contenido:
            # Obtener usuario (autenticado o demo)
            usuario = request.user if request.user.is_authenticated else get_demo_user()
            
            # Crear la nota
            nota_data = {
                'expediente': expediente,
                'contenido': contenido,
                'color': color,
                'posicion_x': posicion_x,
                'posicion_y': posicion_y,
                'creada_por': usuario,
            }
            
            # Si se especifica un documento, asociar la nota
            if documento_id:
                try:
                    documento = DocumentoExpediente.objects.get(pk=documento_id, expediente=expediente)
                    nota_data['documento'] = documento
                except DocumentoExpediente.DoesNotExist:
                    pass
            
            nota = NotaExpediente.objects.create(**nota_data)
            
            if request.headers.get('Accept') == 'application/json':
                return JsonResponse({
                    'success': True,
                    'nota_id': nota.pk,
                    'mensaje': 'Nota creada correctamente'
                })
            else:
                messages.success(request, 'Nota agregada correctamente.')
    
    return redirect('expedientes:visualizador', expediente_id=expediente.pk)


@login_required
def editar_nota(request, expediente_id, nota_id):
    """Vista para editar una nota existente"""
    expediente = get_object_or_404(Expediente, pk=expediente_id)
    nota = get_object_or_404(NotaExpediente, pk=nota_id, expediente=expediente)
    
    if request.method == 'POST':
        contenido = request.POST.get('contenido')
        color = request.POST.get('color', nota.color)
        posicion_x = int(request.POST.get('posicion_x', nota.posicion_x))
        posicion_y = int(request.POST.get('posicion_y', nota.posicion_y))
        
        if contenido:
            nota.contenido = contenido
            nota.color = color
            nota.posicion_x = posicion_x
            nota.posicion_y = posicion_y
            nota.save()
            
            if request.headers.get('Accept') == 'application/json':
                return JsonResponse({
                    'success': True,
                    'mensaje': 'Nota actualizada correctamente'
                })
            else:
                messages.success(request, 'Nota actualizada correctamente.')
    
    return redirect('expedientes:visualizador', expediente_id=expediente.pk)


@login_required
def eliminar_nota(request, expediente_id, nota_id):
    """Vista para eliminar (archivar) una nota"""
    expediente = get_object_or_404(Expediente, pk=expediente_id)
    nota = get_object_or_404(NotaExpediente, pk=nota_id, expediente=expediente)
    
    if request.method == 'POST':
        nota.activa = False
        nota.save()
        
        if request.headers.get('Accept') == 'application/json':
            return JsonResponse({
                'success': True,
                'mensaje': 'Nota eliminada correctamente'
            })
        else:
            messages.success(request, 'Nota eliminada correctamente.')
    
    return redirect('expedientes:visualizador', expediente_id=expediente.pk)


def descargar_expediente_completo(request, expediente_id):
    """Vista para descargar todos los documentos del expediente en un ZIP"""
    expediente = get_object_or_404(Expediente, pk=expediente_id)
    documentos = DocumentoExpediente.objects.filter(expediente=expediente).exclude(archivo='')
    
    if not documentos:
        messages.error(request, 'No hay documentos para descargar en este expediente.')
        return redirect('expedientes:visualizador', expediente_id=expediente.pk)
    
    # Crear respuesta HTTP para el archivo ZIP
    response = HttpResponse(content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="Expediente_{expediente.numero_expediente}.zip"'
    
    # Crear archivo ZIP en memoria
    with zipfile.ZipFile(response, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for i, documento in enumerate(documentos, 1):
            if documento.archivo and os.path.exists(documento.archivo.path):
                # Obtener extensión del archivo original
                _, extension = os.path.splitext(documento.archivo.name)
                
                # Crear nombre del archivo en el ZIP
                nombre_archivo = f"{i:02d}_{documento.nombre_documento}{extension}"
                
                # Agregar archivo al ZIP
                zip_file.write(documento.archivo.path, nombre_archivo)
    
    return response


def convertir_documento_a_pdf(file_path, nombre_documento):
    """Convierte cualquier tipo de documento a PDF y retorna el buffer"""
    file_extension = os.path.splitext(file_path)[1].lower()
    buffer = io.BytesIO()
    
    try:
        if file_extension == '.pdf':
            # PDF: devolver tal como está
            with open(file_path, 'rb') as f:
                buffer.write(f.read())
            buffer.seek(0)
            return buffer
            
        elif file_extension in ['.docx', '.doc']:
            # Word: extraer texto y convertir a PDF
            doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch, bottomMargin=1*inch)
            story = []
            styles = getSampleStyleSheet()
            
            # Título del documento
            story.append(Paragraph(f"<b>{nombre_documento}</b>", styles['Title']))
            story.append(Spacer(1, 20))
            
            try:
                # Leer contenido del Word
                document = Document(file_path)
                for paragraph in document.paragraphs:
                    if paragraph.text.strip():
                        story.append(Paragraph(paragraph.text, styles['Normal']))
                        story.append(Spacer(1, 12))
                        
                if not story or len(story) <= 2:  # Solo título
                    story.append(Paragraph("(Documento Word sin contenido de texto extraíble)", styles['Italic']))
                    
            except Exception as e:
                story.append(Paragraph(f"Error al leer Word: {str(e)}", styles['Normal']))
            
            doc.build(story)
            buffer.seek(0)
            return buffer
            
        elif file_extension in ['.xlsx', '.xls']:
            # Excel: extraer datos y convertir a PDF
            doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch, bottomMargin=1*inch)
            story = []
            styles = getSampleStyleSheet()
            
            # Título del documento
            story.append(Paragraph(f"<b>{nombre_documento}</b>", styles['Title']))
            story.append(Spacer(1, 20))
            
            try:
                workbook = load_workbook(file_path, read_only=True)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    story.append(Paragraph(f"<b>Hoja: {sheet_name}</b>", styles['Heading2']))
                    story.append(Spacer(1, 10))
                    
                    # Leer filas del Excel
                    row_count = 0
                    for row in sheet.iter_rows(values_only=True):
                        if row_count > 50:  # Limitar a 50 filas
                            story.append(Paragraph("... (más filas omitidas)", styles['Italic']))
                            break
                        if any(cell is not None for cell in row):
                            row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                            if row_text.strip():
                                story.append(Paragraph(row_text, styles['Normal']))
                                story.append(Spacer(1, 6))
                        row_count += 1
                    
                    story.append(Spacer(1, 20))
                    
            except Exception as e:
                story.append(Paragraph(f"Error al leer Excel: {str(e)}", styles['Normal']))
            
            doc.build(story)
            buffer.seek(0)
            return buffer
            
        elif file_extension in ['.jpg', '.jpeg', '.png', '.tiff', '.bmp']:
            # Imagen: crear PDF con la imagen
            doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
            story = []
            styles = getSampleStyleSheet()
            
            # Título
            story.append(Paragraph(f"<b>{nombre_documento}</b>", styles['Heading2']))
            story.append(Spacer(1, 20))
            
            try:
                # Agregar imagen
                pil_image = PILImage.open(file_path)
                
                # Calcular tamaño
                page_width = A4[0] - 1*inch
                page_height = A4[1] - 2*inch
                
                img_width, img_height = pil_image.size
                width_ratio = page_width / img_width
                height_ratio = page_height / img_height
                ratio = min(width_ratio, height_ratio, 1)
                
                final_width = img_width * ratio
                final_height = img_height * ratio
                
                img = ReportLabImage(file_path, width=final_width, height=final_height)
                story.append(img)
                
            except Exception as e:
                story.append(Paragraph(f"Error al cargar imagen: {str(e)}", styles['Normal']))
            
            doc.build(story)
            buffer.seek(0)
            return buffer
            
        else:
            # Otros archivos: crear página informativa
            doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch, bottomMargin=1*inch)
            story = []
            styles = getSampleStyleSheet()
            
            story.append(Paragraph(f"<b>{nombre_documento}</b>", styles['Title']))
            story.append(Spacer(1, 20))
            story.append(Paragraph(f"Tipo de archivo: {file_extension.upper()}", styles['Normal']))
            story.append(Paragraph("Este tipo de archivo no se puede convertir automáticamente.", styles['Italic']))
            story.append(Paragraph("Descargue el archivo original para ver su contenido.", styles['Italic']))
            
            doc.build(story)
            buffer.seek(0)
            return buffer
            
    except Exception as e:
        # En caso de error, crear PDF con mensaje de error
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        story.append(Paragraph(f"<b>ERROR: {nombre_documento}</b>", styles['Title']))
        story.append(Spacer(1, 20))
        story.append(Paragraph(f"Error al procesar: {str(e)}", styles['Normal']))
        
        doc.build(story)
        buffer.seek(0)
        return buffer


@login_required
def generar_pdf_completo(request, expediente_id):
    """Vista para convertir TODOS los documentos a PDF y unirlos"""
    expediente = get_object_or_404(Expediente, pk=expediente_id)
    documentos = DocumentoExpediente.objects.filter(
        expediente=expediente
    ).exclude(archivo='').order_by('fecha_subida')
    
    if not documentos:
        messages.error(request, 'No hay documentos en este expediente.')
        return redirect('expedientes:visualizador', expediente_id=expediente.pk)
    
    try:
        # Crear writer para el PDF final
        pdf_writer = PdfWriter()
        documentos_procesados = 0
        
        # Convertir cada documento a PDF y agregarlo
        for documento in documentos:
            if not documento.archivo or not os.path.exists(documento.archivo.path):
                continue
                
            try:
                # Convertir documento a PDF
                pdf_buffer = convertir_documento_a_pdf(
                    documento.archivo.path, 
                    documento.nombre_documento
                )
                
                # Leer el PDF convertido y agregar sus páginas
                pdf_reader = PdfReader(pdf_buffer)
                for page in pdf_reader.pages:
                    pdf_writer.add_page(page)
                
                documentos_procesados += 1
                pdf_buffer.close()
                
            except Exception as e:
                # Si un documento falla, continuar con el siguiente
                continue
        
        if documentos_procesados == 0:
            messages.error(request, 'No se pudo procesar ningún documento.')
            return redirect('expedientes:visualizador', expediente_id=expediente.pk)
        
        # Crear respuesta HTTP
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Expediente_Completo_{expediente.numero_expediente}.pdf"'
        
        # Escribir PDF final
        pdf_writer.write(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error al generar PDF: {str(e)}')
        return redirect('expedientes:visualizador', expediente_id=expediente.pk)


@login_required
def crear_comentario_area(request, expediente_id):
    """Vista para crear comentarios en áreas específicas"""
    from .models import ComentarioArea
    
    expediente = get_object_or_404(Expediente, pk=expediente_id)
    
    if request.method == 'POST':
        etapa = request.POST.get('etapa')
        contenido = request.POST.get('contenido')
        
        if etapa and contenido:
            # Crear el comentario
            comentario = ComentarioArea.objects.create(
                expediente=expediente,
                etapa=etapa,
                usuario=request.user,
                contenido=contenido
            )
            
            # Procesar menciones
            comentario.procesar_menciones()
            
            messages.success(request, 'Comentario agregado exitosamente.')
        else:
            messages.error(request, 'Debe proporcionar el contenido del comentario.')
    
    return redirect('expedientes:visualizador', expediente_id=expediente.pk)


@login_required
def marcar_notificacion_leida(request, notificacion_id):
    """Vista para marcar una notificación como leída"""
    from .models import Notificacion
    
    notificacion = get_object_or_404(Notificacion, pk=notificacion_id, usuario=request.user)
    notificacion.marcar_como_leida()
    
    # Redirigir al enlace de la notificación si existe
    if notificacion.enlace:
        return redirect(notificacion.enlace)
    
    return redirect('expedientes:lista')


@login_required
def obtener_notificaciones(request):
    """Vista AJAX para obtener notificaciones no leídas"""
    from .models import Notificacion
    
    # Obtener todas las notificaciones no leídas para contar el total
    total_no_leidas = Notificacion.objects.filter(
        usuario=request.user,
        leida=False
    ).count()
    
    # Obtener las últimas 10 notificaciones para mostrar
    notificaciones = Notificacion.objects.filter(
        usuario=request.user,
        leida=False
    ).order_by('-fecha_creacion')[:10]
    
    data = []
    for notif in notificaciones:
        data.append({
            'id': notif.id,
            'titulo': notif.titulo,
            'mensaje': notif.mensaje,
            'tipo': notif.tipo,
            'fecha_creacion': notif.fecha_creacion.isoformat(),
            'leida': notif.leida,
            'enlace': notif.enlace,
            'expediente': notif.expediente.numero_expediente if notif.expediente else '',
        })
    
    return JsonResponse({
        'notificaciones': data,
        'total_no_leidas': total_no_leidas
    })


@login_required
def obtener_usuarios_mencion(request):
    """Vista AJAX para autocompletar usuarios en menciones"""
    
    query = request.GET.get('q', '')
    usuarios = User.objects.filter(
        username__icontains=query,
        is_active=True
    ).exclude(id=request.user.id)[:10]
    
    data = [{'username': u.username, 'full_name': u.get_full_name() or u.username} for u in usuarios]
    
    return JsonResponse({'usuarios': data})


def expedientes_por_tipo(request):
    """API endpoint para obtener estadísticas de expedientes por tipo"""
    expedientes_por_tipo = Expediente.objects.values('tipo_expediente').annotate(
        total=Count('id')
    ).order_by('tipo_expediente')
    
    # Mapear nombres de tipos
    tipo_labels = dict(Expediente.TIPO_CHOICES) if hasattr(Expediente, 'TIPO_CHOICES') else {}
    
    data = []
    for item in expedientes_por_tipo:
        data.append({
            'tipo': item['tipo_expediente'],
            'label': tipo_labels.get(item['tipo_expediente'], item['tipo_expediente']),
            'total': item['total']
        })
    
    return JsonResponse({'data': data})


def estadisticas_semanales(request):
    """API endpoint para obtener estadísticas semanales de expedientes"""
    from datetime import timedelta
    
    # Obtener los últimos 7 días
    hoy = timezone.now().date()
    inicio_semana = hoy - timedelta(days=6)
    
    # Generar datos por día
    data = []
    for i in range(7):
        fecha = inicio_semana + timedelta(days=i)
        expedientes_creados = Expediente.objects.filter(
            fecha_creacion__date=fecha
        ).count()
        expedientes_completados = Expediente.objects.filter(
            fecha_actualizacion__date=fecha,
            estado_actual='completado'
        ).count()
        
        data.append({
            'fecha': fecha.strftime('%Y-%m-%d'),
            'dia': fecha.strftime('%a'),
            'creados': expedientes_creados,
            'completados': expedientes_completados
        })
    
    return JsonResponse({'data': data})


@login_required
def eliminar_documento_area(request, documento_id):
    """Vista para eliminar un documento de un área"""
    documento = get_object_or_404(DocumentoExpediente, pk=documento_id)
    expediente = documento.expediente
    
    if request.method == 'POST':
        nombre_doc = documento.nombre_documento
        
        # Eliminar el archivo físico si existe
        if documento.archivo:
            try:
                if os.path.exists(documento.archivo.path):
                    os.remove(documento.archivo.path)
            except Exception:
                pass  # Ignorar errores al eliminar archivo físico
        
        # Eliminar el registro
        documento.delete()
        
        # Registrar en historial
        usuario = request.user if request.user.is_authenticated else get_demo_user()
        HistorialExpediente.objects.create(
            expediente=expediente,
            usuario=usuario,
            accion='Documento eliminado',
            descripcion=f'Documento "{nombre_doc}" eliminado'
        )
        
        messages.success(request, f'Documento "{nombre_doc}" eliminado correctamente.')
        
        if request.headers.get('Accept') == 'application/json':
            return JsonResponse({'success': True, 'mensaje': 'Documento eliminado correctamente'})
    
    return redirect('expedientes:detalle', pk=expediente.pk)


@login_required
def obtener_progreso_documentos(request, expediente_id):
    """Vista para obtener el progreso de documentos por área"""
    from django.db.models import Count, Q, Exists, OuterRef
    
    try:
        # Obtener el expediente
        expediente = Expediente.objects.get(pk=expediente_id)
        
        # Obtener todas las áreas únicas que tienen documentos
        areas_con_documentos = DocumentoExpediente.objects.filter(
            expediente_id=expediente_id
        ).values_list('area_id', flat=True).distinct()
        
        # Obtener el total de áreas únicas en el sistema
        from django.apps import apps
        Area = apps.get_model('digitalizacion', 'Area')
        total_areas = Area.objects.all().count()
        
        # Calcular el progreso
        areas_completadas = len(areas_con_documentos)
        
        if total_areas > 0:
            porcentaje_completado = (areas_completadas / total_areas) * 100
        else:
            porcentaje_completado = 0
        
        # Obtener el estado de cada área
        areas_estado = []
        areas = Area.objects.all()
        
        for area in areas:
            tiene_documentos = area.id in areas_con_documentos
            areas_estado.append({
                'id': area.id,
                'nombre': area.nombre,
                'completada': tiene_documentos,
                'icono': 'bi-check-circle-fill text-success' if tiene_documentos else 'bi-x-circle-fill text-muted'
            })
        
        return JsonResponse({
            'porcentaje_completo': round(porcentaje_completado, 2),
            'areas_completadas': areas_completadas,
            'total_areas': total_areas,
            'areas': areas_estado,
            'status': 'success'
        })
        
    except Expediente.DoesNotExist:
        return JsonResponse({
            'error': 'Expediente no encontrado',
            'status': 'error'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'status': 'error'
        }, status=500)

@login_required
def obtener_documentos_por_area(request, expediente_id, area_id):
    """
    Vista para obtener los documentos de un área específica de un expediente.
    """
    try:
        # Obtener el expediente
        expediente = get_object_or_404(Expediente, pk=expediente_id)
        
        # Obtener el área
        area = get_object_or_404(Area, pk=area_id)
        
        # Obtener los documentos del área
        documentos = DocumentoExpediente.objects.filter(
            expediente=expediente,
            area=area
        ).select_related('subido_por').order_by('-fecha_subida')
        
        # Formatear los documentos
        documentos_data = []
        for doc in documentos:
            documentos_data.append({
                'id': doc.id,
                'nombre_documento': doc.nombre_documento,
                'archivo': os.path.basename(doc.archivo.name) if doc.archivo else None,
                'archivo_url': doc.archivo.url if doc.archivo else None,
                'fecha_subida': doc.fecha_subida.isoformat(),
                'subido_por': doc.subido_por.get_full_name() if doc.subido_por else 'Usuario desconocido',
                'tamanio': doc.archivo.size if doc.archivo else 0,
                'tipo': os.path.splitext(doc.archivo.name)[1].lstrip('.').lower() if doc.archivo else ''
            })
        
        return JsonResponse(documentos_data, safe=False)
        
    except Exception as e:
        import traceback
        print(f"Error en obtener_documentos_por_area: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'error': 'Error al obtener los documentos del área',
            'detalle': str(e)
        }, status=500)
